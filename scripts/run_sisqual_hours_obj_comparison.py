#!/usr/bin/env python3
"""Run the Sisqual MD5 ILP objective variants and write an Excel comparison."""

from __future__ import annotations

import csv
import json
import sys
import time
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
SCHEDULER_SRC = ROOT / "src" / "scheduler"
OPENPYXL_SITE = ROOT / ".venv_excel" / "lib" / "python3.13" / "site-packages"

sys.path.insert(0, str(SCHEDULER_SRC))
if OPENPYXL_SITE.is_dir():
    sys.path.insert(0, str(OPENPYXL_SITE))

import pulp  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from algorithms.ILP_Sisqual_Hours_MathematicalDefinition5 import (  # noqa: E402
    SisqualProblem5ILP,
)


_CBC_CMD = pulp.PULP_CBC_CMD


def _quiet_cbc_cmd(*args, **kwargs):
    kwargs["msg"] = 0
    return _CBC_CMD(*args, **kwargs)


pulp.PULP_CBC_CMD = _quiet_cbc_cmd


RUNS = [
    (
        "SISQUAL_HOURS_OCTOBER_COMPLETE_OBJ1",
        ROOT / "data/problems/SISQUAL_HOURS_OCTOBER_COMPLETE_OBJ1/problem.json",
    ),
    (
        "SISQUAL_HOURS_OCTOBER_COMPLETE_OBJ1_4",
        ROOT / "data/problems/SISQUAL_HOURS_OCTOBER_COMPLETE_OBJ1_4/problem.json",
    ),
    (
        "SISQUAL_HOURS_OCTOBER_COMPLETE_OBJ1_5",
        ROOT / "data/problems/SISQUAL_HOURS_OCTOBER_COMPLETE_OBJ1_5/problem.json",
    ),
    (
        "SISQUAL_HOURS_OCTOBER_COMPLETE_OBJ1_4_5",
        ROOT / "data/problems/SISQUAL_COMPLETE/problem.json",
    ),
]

OUTPUT_DIR = ROOT / "docs" / "comparisons"
SCHEDULE_DIR = ROOT / "shared_tmp" / "obj_compare_20260511"
WORKBOOK_PATH = OUTPUT_DIR / "SISQUAL_HOURS_OCTOBER_MD5_OBJECTIVE_COMPARISON_20260511.xlsx"
RAW_PATH = OUTPUT_DIR / "sisqual_hours_october_md5_objective_comparison_20260511_raw.json"

OFF_MARKERS = {"", "0", "DO", "FDO", "VAC", "NOT", "MED", "OFF", "CLOSED", "UNASSIGNED"}
STAFF_TEAM = "Employees"


def minutes(value: str) -> int | None:
    text = str(value or "").strip()
    if not text:
        return None
    if ":" not in text:
        return int(float(text) * 60)
    hours, mins = text.split(":", 1)
    return int(hours) * 60 + int(mins)


def hhmm(value: int) -> str:
    return f"{value // 60:02d}:{value % 60:02d}"


def slot_starts(start: int, end: int) -> list[int]:
    return list(range(start, end, 30)) if start is not None and end is not None and end > start else []


def parse_segments(cell: str) -> list[dict]:
    text = str(cell or "").strip()
    if not text or text.upper() in OFF_MARKERS:
        return []
    segments = []
    for part in text.split("|"):
        token = part.strip()
        if not token or "@" not in token or "-" not in token:
            continue
        time_range, team = token.rsplit("@", 1)
        start_text, end_text = time_range.split("-", 1)
        start = minutes(start_text)
        end = minutes(end_text)
        if start is not None and end is not None and end > start:
            segments.append({"start": start, "end": end, "team": team.strip()})
    return sorted(segments, key=lambda item: (item["start"], item["end"], item["team"]))


def load_schedule_csv(path: Path) -> tuple[list[str], dict[str, dict[str, list[dict]]]]:
    with path.open(newline="", encoding="utf-8") as file:
        rows = list(csv.reader(file))
    days = rows[0][1:]
    schedule = {}
    for row in rows[1:]:
        employee_id = row[0]
        schedule[employee_id] = {
            day: parse_segments(row[index]) if index < len(row) else []
            for index, day in enumerate(days, start=1)
        }
    return days, schedule


def load_schedule_rules(path: Path) -> dict[str, dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as file:
        reader = csv.DictReader(file)
        return {
            row["employee_id"].strip(): {
                key: str(value or "").strip()
                for key, value in row.items()
                if key != "employee_id"
            }
            for row in reader
            if row.get("employee_id")
        }


def employee_meta(problem: dict) -> dict[str, dict]:
    result = {}
    for raw in problem.get("employees", {}).get("competency", []):
        teams = []
        for index, team in enumerate(raw.get("teams", []), start=1):
            code = str(team.get("code", "")).strip()
            if not code:
                continue
            level = team.get("level", index)
            teams.append({"code": code, "level": float(level)})
        teams.sort(key=lambda item: (item["level"], item["code"]))
        result[str(raw["id"])] = {
            "name": raw.get("name", raw["id"]),
            "primary_team": teams[0]["code"] if teams else None,
            "levels": {item["code"]: item["level"] for item in teams},
        }
    return result


def demanded_minutes(rule: str) -> int | None:
    text = str(rule or "").strip()
    if not text or text.upper() in OFF_MARKERS:
        return None
    if text.upper().startswith("EQUALS:"):
        start_text, end_text = text.split(":", 1)[1].split("-", 1)
        return minutes(end_text) - minutes(start_text)
    try:
        return int(round(float(text) * 60))
    except ValueError:
        return None


def rule_compliant(rule: str, segments: list[dict]) -> bool:
    text = str(rule or "").strip()
    if not text:
        return True
    if text.upper() in OFF_MARKERS:
        return not segments
    total_minutes = sum(max(segment["end"] - segment["start"], 0) for segment in segments)
    expected = demanded_minutes(text)
    if expected is None:
        return not segments
    return total_minutes == expected


def build_coverage(schedule: dict[str, dict[str, list[dict]]]) -> Counter:
    coverage = Counter()
    for by_day in schedule.values():
        for day, segments in by_day.items():
            for segment in segments:
                for slot in slot_starts(segment["start"], segment["end"]):
                    coverage[(day, segment["team"], slot)] += 1
                    if segment["team"] != STAFF_TEAM:
                        coverage[(day, STAFF_TEAM, slot)] += 1
    return coverage


def demand_metrics(problem_path: Path, problem: dict, coverage: Counter) -> tuple[dict, list[dict]]:
    work_periods = {}
    for period in problem["demand"]["workPeriods"]:
        work_periods[period["code"]] = {
            "team": str(period["code"]).split("_", 1)[0].title(),
            "start": minutes(period["timeRange"]["start"]),
            "end": minutes(period["timeRange"]["end"]),
            "name": period.get("name") or period["code"],
        }
    demand_path = problem_path.parent / problem.get("demand", {}).get("dataFile", "demand.csv")
    required_total = fulfilled_total = actual_total = overstaff_total = 0
    critical_underfilled = max_shortage = 0
    team_totals = defaultdict(lambda: {"required": 0, "actual": 0, "fulfilled": 0, "overstaff": 0})
    underfilled = []

    with demand_path.open(newline="", encoding="utf-8") as file:
        for row in csv.DictReader(file):
            period = work_periods[row["workPeriod"]]
            required = int(row["minimum"])
            team = row["team"]
            values = [coverage.get((row["date"], team, slot), 0) for slot in slot_starts(period["start"], period["end"])]
            period_actual = min(values) if values else 0
            shortage = max(required - period_actual, 0)
            if shortage:
                critical_underfilled += sum(1 for actual in values if actual < required)
                max_shortage = max(max_shortage, shortage)
                underfilled.append(
                    {
                        "date": row["date"],
                        "team": team,
                        "workPeriod": row["workPeriod"],
                        "start": hhmm(period["start"]),
                        "end": hhmm(period["end"]),
                        "required": required,
                        "actual": period_actual,
                        "shortage": shortage,
                    }
                )
            for actual in values:
                fulfilled = min(actual, required)
                required_total += required
                actual_total += actual
                fulfilled_total += fulfilled
                team_totals[team]["required"] += required
                team_totals[team]["actual"] += actual
                team_totals[team]["fulfilled"] += fulfilled
                if actual > required:
                    overstaff = actual - required
                    overstaff_total += overstaff
                    team_totals[team]["overstaff"] += overstaff

    team_rows = []
    for team, values in sorted(team_totals.items()):
        required = values["required"]
        fulfilled = values["fulfilled"]
        team_rows.append(
            {
                "team": team,
                "required": required,
                "actual": values["actual"],
                "filled": fulfilled,
                "gap": required - fulfilled,
                "overstaff": values["overstaff"],
                "coverageRate": round(fulfilled / required * 100, 2) if required else 100.0,
            }
        )

    return (
        {
            "weightedMinimumCoverageRate": round(fulfilled_total / required_total * 100, 2) if required_total else 100.0,
            "totalMinimumGap": required_total - fulfilled_total,
            "criticalUnderfilledPeriods": critical_underfilled,
            "maxPeriodShortage": max_shortage,
            "totalOverstaff": overstaff_total,
            "totalActualCoverageSlots": actual_total,
        },
        team_rows,
    )


def assignment_metrics(problem_path: Path, problem: dict, days: list[str], schedule: dict[str, dict[str, list[dict]]]) -> dict:
    rules = load_schedule_rules(problem_path.parent / problem.get("scheduleInput", {}).get("dataFile", "schedule_input.csv"))
    meta = employee_meta(problem)
    total_switches = total_fragmented = 0
    total_worked = total_primary = 0
    compliant = evaluated = hours_match = hours_evaluated = 0
    availability_violations = consecutive_violations = rest_violations = 0
    employee_rows = []
    min_rest_hours = 11.0
    for constraint in problem.get("constraints", {}).get("hard", []):
        if constraint.get("type") == "min_rest_hours" and constraint.get("enabled", True):
            min_rest_hours = float(constraint.get("params", {}).get("hours", min_rest_hours))

    for employee_id in sorted(set(schedule) | set(rules) | set(meta)):
        by_day = schedule.get(employee_id, {})
        employee_rules = rules.get(employee_id, {})
        primary = meta.get(employee_id, {}).get("primary_team")
        worked_minutes = primary_minutes = 0
        team_switches = fragmented = employee_availability = employee_consecutive = employee_rest = 0
        streak = 0
        previous_end = None
        previous_date = None
        team_minutes = defaultdict(int)

        for day in days:
            segments = by_day.get(day, [])
            worked = bool(segments)
            day_minutes = sum(max(segment["end"] - segment["start"], 0) for segment in segments)
            worked_minutes += day_minutes
            primary_minutes += sum(
                max(segment["end"] - segment["start"], 0)
                for segment in segments
                if primary is None or segment["team"] == primary
            )
            for segment in segments:
                team_minutes[segment["team"]] += max(segment["end"] - segment["start"], 0)
            if len(segments) > 1:
                fragmented += 1
            for previous, current in zip(segments, segments[1:]):
                if previous["team"] != current["team"]:
                    team_switches += 1
            if day in employee_rules:
                evaluated += 1
                if rule_compliant(employee_rules[day], segments):
                    compliant += 1
                else:
                    availability_violations += 1
                    employee_availability += 1
                expected = demanded_minutes(employee_rules[day])
                if expected is not None:
                    hours_evaluated += 1
                    if day_minutes == expected:
                        hours_match += 1
            if worked:
                streak += 1
                if streak > 5:
                    consecutive_violations += 1
                    employee_consecutive += 1
            else:
                streak = 0
            current_date = date.fromisoformat(day)
            if worked and previous_end is not None and previous_date == current_date - timedelta(days=1):
                rest_hours = ((24 * 60 - previous_end) + segments[0]["start"]) / 60.0
                if rest_hours < min_rest_hours:
                    rest_violations += 1
                    employee_rest += 1
            if worked:
                previous_end = max(segment["end"] for segment in segments)
                previous_date = current_date
            else:
                previous_end = None
                previous_date = None

        total_switches += team_switches
        total_fragmented += fragmented
        total_worked += worked_minutes
        total_primary += primary_minutes
        employee_rows.append(
            {
                "employeeId": employee_id,
                "name": meta.get(employee_id, {}).get("name", employee_id),
                "workedHours": round(worked_minutes / 60, 2),
                "primaryTeam": primary,
                "primaryTeamUtilizationRate": round(primary_minutes / worked_minutes * 100, 2) if worked_minutes else 100.0,
                "teamSwitches": team_switches,
                "fragmentedWorkDays": fragmented,
                "availabilityViolations": employee_availability,
                "consecutiveDaysViolations": employee_consecutive,
                "minRestViolations": employee_rest,
                "teamHours": {team: round(value / 60, 2) for team, value in sorted(team_minutes.items())},
            }
        )

    employee_rows.sort(key=lambda item: (-item["availabilityViolations"], -item["teamSwitches"], item["employeeId"]))
    return {
        "intraDayTeamSwitches": total_switches,
        "fragmentedWorkDays": total_fragmented,
        "primaryTeamUtilizationRate": round(total_primary / total_worked * 100, 2) if total_worked else 100.0,
        "nonPrimaryTeamHours": round((total_worked - total_primary) / 60, 2),
        "totalWorkedHours": round(total_worked / 60, 2),
        "durationComplianceRate": round(compliant / evaluated * 100, 2) if evaluated else 100.0,
        "demandedHoursComplianceRate": round(hours_match / hours_evaluated * 100, 2) if hours_evaluated else 100.0,
        "availabilityViolations": availability_violations,
        "consecutiveDaysViolations": consecutive_violations,
        "minRestViolations": rest_violations,
        "employeeRows": employee_rows,
    }


def variable_sum(variables: dict) -> float:
    return float(sum((pulp.value(variable) or 0.0) for variable in variables.values()))


def run_case(title: str, problem_path: Path) -> dict:
    with problem_path.open(encoding="utf-8") as file:
        problem = json.load(file)

    scheduler = SisqualProblem5ILP(str(problem_path), max_time_minutes="10")
    scheduler.build_model()
    started = time.perf_counter()
    status_code = scheduler.solve(gap_rel=0.01)
    elapsed = time.perf_counter() - started
    status = pulp.LpStatus.get(status_code, "Unknown")

    csv_path = SCHEDULE_DIR / f"{title}.csv"
    scheduler.export_csv(str(csv_path))
    days, schedule = load_schedule_csv(csv_path)
    coverage = build_coverage(schedule)
    minimums, team_rows = demand_metrics(problem_path, problem, coverage)
    assignments = assignment_metrics(problem_path, problem, days, schedule)

    return {
        "title": title,
        "problemId": problem.get("metadata", {}).get("problemId"),
        "problemPath": str(problem_path.relative_to(ROOT)),
        "csvPath": str(csv_path.relative_to(ROOT)),
        "status": status,
        "elapsedSeconds": round(elapsed, 3),
        "objectiveValue": scheduler.objective_value,
        "switchObjectiveValue": scheduler.switch_objective_value,
        "objectiveWeights": {
            "obj1": scheduler.objective1_weight,
            "obj4": scheduler.objective4_weight,
            "obj5": scheduler.objective5_weight,
            "skillSwitch": scheduler.skill_switch_weight,
        },
        "modelSize": {
            "variables": len(scheduler.model.variables()),
            "constraints": len(scheduler.model.constraints),
            "x": len(scheduler.x),
            "y": len(scheduler.y),
            "switch": len(scheduler.switch),
        },
        "componentValues": {
            "shortageSlots": variable_sum(scheduler.shortage),
            "shortageUnits": variable_sum(scheduler.shortage_unit),
            "preferredDayOffWorked": variable_sum(scheduler.preferred_day_work),
            "skillSwitches": variable_sum(scheduler.switch),
        },
        "minimums": minimums,
        "teamBreakdown": team_rows,
        "assignments": assignments,
    }


def append_rows(ws, rows):
    for row in rows:
        ws.append(row)


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 42)
        ws.column_dimensions[get_column_letter(column[0].column)].width = width


def write_workbook(results: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    append_rows(
        ws,
        [
            [
                "Problem",
                "Status",
                "Time (s)",
                "Primary Objective",
                "Switch Objective",
                "Obj1 Weight",
                "Obj4 Weight",
                "Obj5 Weight",
                "Coverage %",
                "Minimum Gap",
                "Critical Underfilled Slots",
                "Max Shortage",
                "Overstaff Slots",
                "Preferred DO Worked",
                "Team Switches",
                "Fragmented Days",
                "Primary Team %",
                "Non-primary Hours",
                "Total Worked Hours",
                "Duration Compliance %",
                "Demanded Hours Compliance %",
                "Availability Violations",
                "Consecutive Violations",
                "Min Rest Violations",
                "Variables",
                "Constraints",
                "Schedule CSV",
                "Problem Path",
            ]
        ],
    )
    for result in results:
        append_rows(
            ws,
            [
                [
                    result["title"],
                    result["status"],
                    result["elapsedSeconds"],
                    result["objectiveValue"],
                    result["switchObjectiveValue"],
                    result["objectiveWeights"]["obj1"],
                    result["objectiveWeights"]["obj4"],
                    result["objectiveWeights"]["obj5"],
                    result["minimums"]["weightedMinimumCoverageRate"],
                    result["minimums"]["totalMinimumGap"],
                    result["minimums"]["criticalUnderfilledPeriods"],
                    result["minimums"]["maxPeriodShortage"],
                    result["minimums"]["totalOverstaff"],
                    result["componentValues"]["preferredDayOffWorked"],
                    result["assignments"]["intraDayTeamSwitches"],
                    result["assignments"]["fragmentedWorkDays"],
                    result["assignments"]["primaryTeamUtilizationRate"],
                    result["assignments"]["nonPrimaryTeamHours"],
                    result["assignments"]["totalWorkedHours"],
                    result["assignments"]["durationComplianceRate"],
                    result["assignments"]["demandedHoursComplianceRate"],
                    result["assignments"]["availabilityViolations"],
                    result["assignments"]["consecutiveDaysViolations"],
                    result["assignments"]["minRestViolations"],
                    result["modelSize"]["variables"],
                    result["modelSize"]["constraints"],
                    result["csvPath"],
                    result["problemPath"],
                ]
            ],
        )
    style_sheet(ws)

    ws = wb.create_sheet("Team Minimums")
    append_rows(ws, [["Problem", "Team", "Required Slots", "Actual Slots", "Filled Slots", "Gap", "Overstaff", "Coverage %"]])
    for result in results:
        for row in result["teamBreakdown"]:
            append_rows(ws, [[result["title"], row["team"], row["required"], row["actual"], row["filled"], row["gap"], row["overstaff"], row["coverageRate"]]])
    style_sheet(ws)

    ws = wb.create_sheet("Employees")
    append_rows(ws, [["Problem", "Employee", "Name", "Worked Hours", "Primary Team", "Primary Team %", "Team Switches", "Fragmented Days", "Availability Violations", "Consecutive Violations", "Min Rest Violations", "Team Hours"]])
    for result in results:
        for row in result["assignments"]["employeeRows"]:
            append_rows(
                ws,
                [[
                    result["title"],
                    row["employeeId"],
                    row["name"],
                    row["workedHours"],
                    row["primaryTeam"],
                    row["primaryTeamUtilizationRate"],
                    row["teamSwitches"],
                    row["fragmentedWorkDays"],
                    row["availabilityViolations"],
                    row["consecutiveDaysViolations"],
                    row["minRestViolations"],
                    json.dumps(row["teamHours"], sort_keys=True),
                ]],
            )
    style_sheet(ws)

    ws = wb.create_sheet("Model")
    append_rows(ws, [["Problem", "Variables", "Constraints", "x Variables", "y Variables", "Switch Variables", "Shortage Slots", "Shortage Units", "Skill Switches"]])
    for result in results:
        append_rows(
            ws,
            [[
                result["title"],
                result["modelSize"]["variables"],
                result["modelSize"]["constraints"],
                result["modelSize"]["x"],
                result["modelSize"]["y"],
                result["modelSize"]["switch"],
                result["componentValues"]["shortageSlots"],
                result["componentValues"]["shortageUnits"],
                result["componentValues"]["skillSwitches"],
            ]],
        )
    style_sheet(ws)

    wb.save(WORKBOOK_PATH)


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    SCHEDULE_DIR.mkdir(parents=True, exist_ok=True)
    results = []
    for title, path in RUNS:
        print(f"[RUN] {title} -> {path.relative_to(ROOT)}", flush=True)
        result = run_case(title, path)
        results.append(result)
        print(
            f"[DONE] {title}: {result['status']} in {result['elapsedSeconds']}s, "
            f"coverage={result['minimums']['weightedMinimumCoverageRate']}%, "
            f"gap={result['minimums']['totalMinimumGap']}",
            flush=True,
        )
    RAW_PATH.write_text(json.dumps(results, indent=2), encoding="utf-8")
    write_workbook(results)
    print(f"[WROTE] {WORKBOOK_PATH.relative_to(ROOT)}")
    print(f"[WROTE] {RAW_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
